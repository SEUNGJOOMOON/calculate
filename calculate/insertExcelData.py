import sys
import os, sys
sys.path.append(os.path.dirname())
from openpyxl import load_workbook
import math
import calculator

# 업로드할 파일 경로
file_path = "/path/data.xlsx"

#core 생성
calc = calculator.Calculator(2203, 'REAL')

print("file name is..." + file_path)
if input("Do you want to continue?") != "y":
    exit()

print("reading logs...")

work_book = load_workbook(file_path, data_only=True)

print("End loading...")

work_sheet = work_book['Sheet1']

sheet_count = work_sheet.max_row + 1

for i in range(2, sheet_count):


    adate = work_sheet.cell(row=i, column=0).value #정산월
    company_id = work_sheet.cell(row=i, column=1).value #회사 코드
    contract_id = work_sheet.cell(row=i, column=2).value #계약 코드
    rate = work_sheet.cell(row=i, column=3).value #정산 요율
    item_code = work_sheet.cell(row=i, column=4).value #상품 코드
    item_price = work_sheet.cell(row=i, column=5).value #상품 가격
    item_sales_cnt = work_sheet.cell(row=i, column=6).value #상품 판매 갯수
    item_return_cnt = work_sheet.cell(row=i, column=7).value #상품 반품 갯수

    if not item_price or math.isnan(item_price):
        item_price = 0
    if not item_sales_cnt or math.isnan(item_sales_cnt):
        item_sales_cnt = 0
    if not item_return_cnt or math.isnan(item_return_cnt):
        item_return_cnt = 0

    sql = """
        INSERT INTO ITEM_SALES
				(adate, company_id, contract_id, rate, item_code, item_price, item_sales_cnt, item_return_cnt)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """

    bind = [adate, company_id, contract_id, rate, item_code, item_price, item_sales_cnt, item_return_cnt]

    calc.setQueryUpdate(sql, bind)

    print(str(i - 1) + "/" + str(sheet_count - 2))

print("Done.")
    


