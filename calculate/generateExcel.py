import sys
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side 
import datetime
sys.path.append(os.path.dirname())
import calculator


calc = calculator.Calculator(2203, 'REAL')

print("adate is..." + str(calc.adate))
if input("Do you want to continue?") != "y":
    exit()


workbook = Workbook()
worksheet = workbook.active
worksheet.title = "20" + str(calc.adate) + "_정산내역"

worksheet.column_dimensions['A'].width = 9
worksheet.column_dimensions['B'].width = 40
worksheet.column_dimensions['C'].width = 40
worksheet.column_dimensions['D'].width = 40

#엑셀 header setting
worksheet.append(['정산월', '상품명', '상품가격', '정산액'])

#excel 스타일 적용
for i in range(1, 4):
    worksheet.cell(row=1, column=i).alignment = Alignment(horizontal='center') # 가운데 정렬
    worksheet.cell(row=1, column=i).border = Border(left=Side(style='thin', color='000000'),right=Side(style='thin', color='000000'),top=Side(style='thin', color='000000'),bottom=Side(style='thin', color='000000'))
    worksheet.cell(row=1, column=i).fill = PatternFill(bgColor="F6F6F6", fgColor="F6F6F6", fill_type="solid")


sql = """
        SELECT 
            id,
            adate,
            company_id,
            contract_id,
            rate,
            item_code,
            item_price,
            item_sales_cnt,
            item_return_cnt
        FROM ITEM_SALES
        WHERE ADATE = %s
    """

bind = [calc.adate]

details = calc.getQueryResult(sql, bind)

for detail in details:

    id = detail["id"] #id
    company_id = detail["company_id"] #회사 코드
    contract_id = detail["contract_id"] #계약 코드
    rate = detail["rate"] #정산 요율


    #excel data setting
    worksheet.append([id, company_id, contract_id, rate])

#create directory
directory = "/path/" + str(calc.adate) + "/"

if not os.path.exists(directory):
    os.makedirs(directory)

fileName = calc.getEncryptedFileName(company_id)

workbook.save(directory + fileName)
print("Done.")